import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from typing import Dict, List, Any
import datetime

class ExcelReporter:
    """Generates professional, color-coded Excel reports based on reconciliation results with visualization."""
    
    # Colors for highlighting
    MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green
    MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red
    MISSING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Light Yellow
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")    # Dark Blue
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    
    def generate_report(self, recon_data: Dict[str, Any], output_path: str):
        """
        Takes the dictionary from ReconEngine and writes a formatted Excel file with charts.
        """
        summary = recon_data.get("summary", {})
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1. Generate Summary Sheet
            self._write_summary_sheet(writer, summary)
            
            # 2. Generate Detailed Deltas Sheet
            self._write_detailed_sheet(writer, recon_data)
            
            # 3. Add Charts (requires direct sheet access)
            self._add_visualizations(writer, summary)

    def _write_summary_sheet(self, writer, summary):
        """Creates a professional dashboard-style summary sheet."""
        total_matched = summary.get("matched", 0) - summary.get("mismatches", 0)
        
        summary_data = [
            ["Reconciliation Summary Report", ""],
            ["Generated At:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Dataset Overview", ""],
            ["Total Rows (Source A)", summary.get("total_a")],
            ["Total Rows (Source B)", summary.get("total_b")],
            ["", ""],
            ["Reconciliation Results", ""],
            ["Perfect Matches", total_matched],
            ["Data Mismatches", summary.get("mismatches")],
            ["Missing in Source B (A only)", len(summary.get("only_in_a", []))],
            ["Missing in Source A (B only)", len(summary.get("only_in_b", []))],
            ["", ""],
            ["Match Rate", f"{(total_matched / summary.get('total_a', 1) * 100):.2f}%"]
        ]
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary Dashboard", index=False, header=False)
        
        ws = writer.sheets["Summary Dashboard"]
        
        # Style the summary
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # Header Styling
        ws['A1'].font = Font(size=16, bold=True, color="4472C4")
        
        # Section Headers
        for row in [4, 8]:
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Add Data Reconciliation Logic explanation
        ws['A22'] = "Data Reconciliation Logic"
        ws['A22'].font = Font(bold=True, size=12)
        ws['A23'] = "1. Keys from both datasets are aligned to identify row presence (Missing vs Found)."
        ws['A24'] = "2. Cell-level comparison is performed only for matched keys based on semantic mapping."
        ws['A25'] = "3. Status 'MISMATCH' indicates the same key exists in both, but values differ."
        ws['A26'] = "4. Status 'ONLY IN A/B' indicates the record is entirely missing from one source."

    def _write_detailed_sheet(self, writer, recon_data):
        """Helper to write and style the detailed reconciliation sheet."""
        summary = recon_data.get("summary", {})
        details = recon_data.get("detail", [])
        
        rows = []
        # Add Mismatches
        for item in details:
            key = item["key"]
            for col, diff in item["differences"].items():
                rows.append({
                    "Unique Key": key,
                    "Field": col,
                    "Source A Value": diff["val_a"],
                    "Source B Value": diff["val_b"],
                    "Difference": "Value Mismatch",
                    "Status": "MISMATCH"
                })
        
        # Add Missing Rows
        for key in summary.get("only_in_a", []):
            rows.append({"Unique Key": key, "Difference": "Missing in B", "Status": "ONLY IN A"})
        for key in summary.get("only_in_b", []):
            rows.append({"Unique Key": key, "Difference": "Missing in A", "Status": "ONLY IN B"})
            
        df_details = pd.DataFrame(rows)
        if df_details.empty:
            df_details = pd.DataFrame(columns=["Unique Key", "Field", "Source A Value", "Source B Value", "Difference", "Status"])
            df_details.loc[0] = ["No differences found", "-", "-", "-", "-", "MATCHED"]

        df_details.to_excel(writer, sheet_name="Reconciliation Details", index=False)
        
        ws = writer.sheets["Reconciliation Details"]
        
        # Apply header styling
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.WHITE_FONT
            cell.alignment = Alignment(horizontal="center")

        # Apply conditional formatting based on Status
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=6).value
            fill = None
            if status == "MISMATCH":
                fill = self.MISMATCH_FILL
            elif status and "ONLY IN" in str(status):
                fill = self.MISSING_FILL
            
            if fill:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = fill

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

    def _add_visualizations(self, writer, summary):
        """Adds charts to the Summary sheet."""
        ws = writer.sheets["Summary Dashboard"]
        
        # Data for the pie chart
        total_matched = summary.get("matched", 0) - summary.get("mismatches", 0)
        mismatches = summary.get("mismatches", 0)
        only_a = len(summary.get("only_in_a", []))
        only_b = len(summary.get("only_in_b", []))
        
        # Place chart data in a hidden area or specific range
        ws["Z1"] = "Category"
        ws["AA1"] = "Count"
        ws["Z2"] = "Matches"
        ws["AA2"] = total_matched
        ws["Z3"] = "Mismatches"
        ws["AA3"] = mismatches
        ws["Z4"] = "Missing in B"
        ws["AA4"] = only_a
        ws["Z5"] = "Missing in A"
        ws["AA5"] = only_b

        # Create Pie Chart
        pie = PieChart()
        labels = Reference(ws, min_col=26, min_row=2, max_row=5)
        data = Reference(ws, min_col=27, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Reconciliation Breakdown"
        pie.width = 10
        pie.height = 7
        
        # Place Pie Chart to the right of the summary table
        ws.add_chart(pie, "D2")

        # Create Bar Chart for Source Comparison
        bar = BarChart()
        ws["Z7"] = "Source"
        ws["AA7"] = "Count"
        ws["Z8"] = "Source A"
        ws["AA8"] = summary.get("total_a")
        ws["Z9"] = "Source B"
        ws["AA9"] = summary.get("total_b")
        
        data_bar = Reference(ws, min_col=27, min_row=7, max_row=9)
        cats_bar = Reference(ws, min_col=26, min_row=8, max_row=9)
        bar.add_data(data_bar, titles_from_data=True)
        bar.set_categories(cats_bar)
        bar.title = "Dataset Size Comparison"
        bar.legend = None
        bar.width = 10
        bar.height = 7
        
        # Place Bar Chart below the Pie Chart, or to the right of the logic section
        ws.add_chart(bar, "M2")

